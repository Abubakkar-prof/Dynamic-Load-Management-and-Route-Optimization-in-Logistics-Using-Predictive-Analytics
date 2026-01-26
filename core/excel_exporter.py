"""
Excel Export Service
Exports data to professionally formatted Excel files
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import io

class ExcelExporter:
    """
    Exports various data to Excel with professional formatting
    """
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_orders(self, orders, filename='orders_export.xlsx'):
        """
        Export orders to Excel with formatting
        
        Args:
            orders: QuerySet or list of Order objects
            filename: Output filename
        """
        # Prepare data
        data = []
        for order in orders:
            data.append({
                'Order ID': order.order_id,
                'Customer Name': order.customer_name,
                'Delivery Address': order.delivery_address,
                'Weight (kg)': order.weight_kg,
                'Volume (m³)': order.volume_m3,
                'Status': order.status,
                'Priority': getattr(order, 'priority', 'Medium'),
                'Time Window Start': f"{order.time_window_start}:00",
                'Time Window End': f"{order.time_window_end}:00",
                'Service Time (min)': order.service_time,
                'Created At': order.created_at.strftime('%Y-%m-%d %H:%M') if hasattr(order, 'created_at') else 'N/A'
            })
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Create Excel writer
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Orders', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Orders']
            
            # Format header row
            for cell in worksheet[1]:
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.border
            
            # Format data rows
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = self.border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Color code status
                    if cell.column == 6:  # Status column
                        if cell.value == 'Delivered':
                            cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                            cell.font = Font(color="065F46", bold=True)
                        elif cell.value == 'In Transit':
                            cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                            cell.font = Font(color="1E40AF", bold=True)
                        elif cell.value == 'Pending':
                            cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                            cell.font = Font(color="92400E", bold=True)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return filename
    
    def export_vehicles(self, vehicles, filename='vehicles_export.xlsx'):
        """
        Export vehicles to Excel
        """
        data = []
        for vehicle in vehicles:
            data.append({
                'Vehicle ID': vehicle.vehicle_id,
                'Type': vehicle.type,
                'Capacity (kg)': vehicle.capacity_kg,
                'Capacity (m³)': vehicle.capacity_vol,
                'Status': vehicle.status,
                'Current Location': f"{vehicle.current_location_lat}, {vehicle.current_location_lon}",
                'Last Updated': vehicle.updated_at.strftime('%Y-%m-%d %H:%M') if hasattr(vehicle, 'updated_at') else 'N/A'
            })
        
        df = pd.DataFrame(data)
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Vehicles', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Vehicles']
            
            # Format header
            for cell in worksheet[1]:
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.border
            
            # Format data
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = self.border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Auto-adjust columns
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 40)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return filename
    
    def export_routes(self, routes, filename='routes_export.xlsx'):
        """
        Export routes to Excel
        """
        data = []
        for route in routes:
            data.append({
                'Route ID': getattr(route, 'id', 'N/A'),
                'Vehicle ID': route.get('vehicle_id', 'N/A'),
                'Total Distance (km)': f"{route.get('total_distance_m', 0) / 1000:.2f}",
                'Total Load (kg)': route.get('total_load_kg', 0),
                'Utilization (%)': route.get('utilization_pct', 0),
                'Number of Stops': len(route.get('route', [])),
                'Status': route.get('status', 'Planned')
            })
        
        df = pd.DataFrame(data)
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Routes', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Routes']
            
            # Format header
            for cell in worksheet[1]:
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.border
            
            # Format data
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = self.border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust columns
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return filename
    
    def export_analytics(self, analytics_data, filename='analytics_export.xlsx'):
        """
        Export analytics dashboard to Excel with multiple sheets
        """
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = {
                'Metric': [
                    'Total Orders',
                    'Delivered Orders',
                    'Pending Orders',
                    'Total Vehicles',
                    'Active Vehicles',
                    'Total Distance (km)',
                    'Average Delivery Time (hrs)',
                    'Fleet Utilization (%)'
                ],
                'Value': [
                    analytics_data.get('total_orders', 0),
                    analytics_data.get('delivered_orders', 0),
                    analytics_data.get('pending_orders', 0),
                    analytics_data.get('total_vehicles', 0),
                    analytics_data.get('active_vehicles', 0),
                    f"{analytics_data.get('total_distance', 0):.2f}",
                    f"{analytics_data.get('avg_delivery_time', 0):.1f}",
                    f"{analytics_data.get('fleet_utilization', 0):.1f}"
                ]
            }
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format summary sheet
            worksheet = writer.sheets['Summary']
            for cell in worksheet[1]:
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            worksheet.column_dimensions['A'].width = 30
            worksheet.column_dimensions['B'].width = 20
        
        return filename
    
    def export_to_buffer(self, data, sheet_name='Data'):
        """
        Export data to Excel buffer (for HTTP response)
        Returns BytesIO buffer
        """
        buffer = io.BytesIO()
        
        df = pd.DataFrame(data)
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Format header
            for cell in worksheet[1]:
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        buffer.seek(0)
        return buffer
